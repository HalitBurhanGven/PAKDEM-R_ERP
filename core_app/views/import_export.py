from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core_app.models import Stock
from core_app.services.stock_import_service import StockImportError, import_stock_rows, parse_stock_workbook
from core_app.stock_forms import StockImportForm


def export_stocks_xlsx(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "stocks"
    headers = ["name", "category", "unit", "sku", "quantity", "created_at"]
    worksheet.append(headers)

    for stock in Stock.objects.order_by("id"):
        worksheet.append([
            stock.name,
            stock.category,
            stock.unit,
            stock.sku,
            stock.quantity,
            stock.created_at.strftime("%d.%m.%Y %H:%M") if stock.created_at else "",
        ])

    for index, _ in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = 18

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="stocks.xlsx"'
    workbook.save(response)
    return response


def import_stocks_xlsx(request):
    if request.method == "POST":
        form = StockImportForm(request.POST, request.FILES)
        if form.is_valid():
            action = (request.POST.get("action") or "preview").strip()
            try:
                parsed = parse_stock_workbook(form.cleaned_data["file"])
            except StockImportError as exc:
                messages.error(request, str(exc))
                return render(request, "core_app/stock_import.html", {"form": form})

            if parsed["missing_headers"]:
                messages.error(request, "Eksik başlıklar: " + ", ".join(parsed["missing_headers"]))
                return render(request, "core_app/stock_import.html", {
                    "form": form,
                    "preview": parsed,
                })

            if action == "preview":
                return render(request, "core_app/stock_import.html", {
                    "form": form,
                    "preview": parsed,
                })

            summary = import_stock_rows(parsed)

            messages.success(
                request,
                f"Yükleme tamam: Yeni={summary.created}, Güncellendi={summary.updated}, Atlandı={summary.skipped}",
            )
            return render(request, "core_app/stock_import.html", {
                "form": StockImportForm(),
                "import_result": {
                    "created": summary.created,
                    "updated": summary.updated,
                    "skipped": summary.skipped,
                    "total": summary.total,
                    "processed": summary.processed,
                    "errors": summary.errors,
                    "error_reasons": summary.error_reasons,
                },
                "preview": parsed,
            })
    else:
        form = StockImportForm()

    return render(request, "core_app/stock_import.html", {"form": form})
