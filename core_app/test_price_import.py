from datetime import date
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from core_app.models import PriceItem, PriceList
from core_app.services.price_import import (
    PriceImportError,
    detect_list_date,
    detect_profile_measure_header_row,
    find_sheet_name,
    import_price_list_workbook,
    parse_profile_measure_sheet,
    parse_hirdavat2_sheet,
    parse_standard_sheet,
    to_decimal,
)


class PriceImportServiceTests(TestCase):
    def test_to_decimal_handles_currency_symbols_and_separators(self):
        self.assertEqual(to_decimal("₺1.234,50"), Decimal("1234.50"))
        self.assertEqual(to_decimal("TRY 99,90"), Decimal("99.90"))
        self.assertEqual(to_decimal(15), Decimal("15.00"))

    def test_detect_list_date_reads_first_10_rows(self):
        wb = Workbook()
        ws = wb.active
        ws["M1"] = date(2026, 3, 22)

        self.assertEqual(detect_list_date(ws), date(2026, 3, 22))

    def test_parse_standard_sheet_extracts_grouped_items(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "HIRDAVAT"
        ws["A2"] = "Pul"
        ws["B2"] = "12,50"
        ws["D1"] = "BOYA"
        ws["D2"] = "Astar"
        ws["E2"] = 25

        items = parse_standard_sheet(ws)

        self.assertEqual(len(items), 2)
        self.assertEqual(items[0].group, "HIRDAVAT")
        self.assertEqual(items[0].name, "Pul")
        self.assertEqual(items[0].price, Decimal("12.50"))
        self.assertEqual(items[1].group, "BOYA")

    def test_parse_hirdavat2_sheet_extracts_items_only_after_group_headers(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "KILITLER"
        ws["B1"] = "FİYAT"
        ws["A2"] = "Barelli Kilit"
        ws["B2"] = "100,00"
        ws["F1"] = "TOPLAM"
        ws["G1"] = "FİYAT"
        ws["F2"] = "Ignore me"
        ws["G2"] = "50,00"

        items = parse_hirdavat2_sheet(ws)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].group, "KILITLER")
        self.assertEqual(items[0].name, "Barelli Kilit")
        self.assertEqual(items[0].price, Decimal("100.00"))

    def test_parse_standard_sheet_skips_nakit_and_toplam_rows(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "NEVSAC"
        ws["A2"] = "40x80x2,40"
        ws["B2"] = "1035,96"
        ws["A3"] = "NAKİT"
        ws["B3"] = "167,22"
        ws["A4"] = "TOPLAM"
        ws["B4"] = "131,51853000000003"

        items = parse_standard_sheet(ws)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].name, "40x80x2,40")
        self.assertEqual(items[0].price, Decimal("1035.96"))

    def test_detect_profile_measure_header_row_finds_cinsi_olcu_layout(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "CİNSİ"
        ws["B1"] = "ÖLÇÜ"
        ws["C1"] = "CİNSİ"
        ws["D1"] = "ANA FİYAT"

        self.assertEqual(detect_profile_measure_header_row(ws), 1)

    def test_parse_profile_measure_sheet_uses_only_cinsi_column_next_to_ana_fiyat(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "CİNSİ"
        ws["B1"] = "ÖLÇÜ"
        ws["C1"] = "CİNSİ"
        ws["D1"] = "ANA FİYAT"
        ws["A2"] = "10x20x1,20"
        ws["B2"] = 35
        ws["C2"] = ""
        ws["D2"] = "167,22"
        ws["A3"] = "10x20x1,50"
        ws["B3"] = 35
        ws["D3"] = "207,00"
        ws["A4"] = ""
        ws["B4"] = 35
        ws["C4"] = "20x20x0,80"
        ws["D4"] = "147,54"
        ws["A5"] = ""
        ws["B5"] = 35
        ws["C5"] = "NAKİT"
        ws["D5"] = "207,00"
        ws["A6"] = ""
        ws["B6"] = ""
        ws["C6"] = "TOPLAM"
        ws["D6"] = "162,80550000000005"

        items = parse_profile_measure_sheet(ws, 1)

        self.assertEqual(len(items), 3)
        self.assertEqual(items[0].name, "10x20x1,20")
        self.assertEqual(items[0].price, Decimal("167.22"))
        self.assertEqual(items[1].name, "10x20x1,50")
        self.assertEqual(items[1].price, Decimal("207.00"))
        self.assertEqual(items[2].name, "20x20x0,80")
        self.assertEqual(items[2].price, Decimal("147.54"))

    def test_parse_standard_sheet_switches_to_profile_measure_layout_when_detected(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "CİNSİ"
        ws["B1"] = "ÖLÇÜ"
        ws["C1"] = "CİNSİ"
        ws["D1"] = "ANA FİYAT"
        ws["A2"] = "10x20x1,20"
        ws["C2"] = "30x30x1,00"
        ws["D2"] = "270,42"

        items = parse_standard_sheet(ws)

        self.assertEqual(len(items), 2)
        self.assertEqual(items[0].name, "10x20x1,20")
        self.assertEqual(items[0].price, Decimal("270.42"))
        self.assertEqual(items[1].name, "30x30x1,00")
        self.assertEqual(items[1].price, Decimal("270.42"))

    def test_import_price_list_workbook_imports_standard_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "HIRDAVAT"
        ws["A1"] = "GRUP A"
        ws["A2"] = "Vida"
        ws["B2"] = "10,50"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        result = import_price_list_workbook(file_obj, "HIRDAVAT", "Liste 1")

        self.assertEqual(result.created_count, 1)
        self.assertEqual(PriceList.objects.count(), 1)
        item = PriceItem.objects.get(price_list=result.price_list)
        self.assertEqual(item.group, "GRUP A")
        self.assertEqual(item.name, "Vida")
        self.assertEqual(item.price, Decimal("10.50"))

    def test_find_sheet_name_matches_turkish_character_variants(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PROFİL10 İSK"

        self.assertEqual(find_sheet_name(wb, "PROFIL10 ISK"), "PROFİL10 İSK")
        self.assertEqual(find_sheet_name(wb, "profil10 isk"), "PROFİL10 İSK")

    def test_import_price_list_workbook_accepts_sheet_name_without_turkish_upper_i(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PROFİL10 İSK"
        ws["A1"] = "GRUP A"
        ws["A2"] = "30x40x1,5"
        ws["B2"] = "10,50"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        result = import_price_list_workbook(file_obj, "PROFIL10 ISK", "Profil Liste")

        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.price_list.sheet_name, "PROFİL10 İSK")

    def test_import_price_list_workbook_rejects_missing_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "FARKLI"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        with self.assertRaises(PriceImportError):
            import_price_list_workbook(file_obj, "HIRDAVAT", "Liste 1")

    def test_import_price_list_workbook_rejects_empty_parse_result(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "HIRDAVAT"
        ws["A1"] = "Sadece Baslik"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        with self.assertRaises(PriceImportError):
            import_price_list_workbook(file_obj, "HIRDAVAT", "Liste 1")

    def test_price_import_preview_hides_price_column(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "HIRDAVAT"
        ws["A1"] = "GRUP A"
        ws["A2"] = "Vida"
        ws["B2"] = "10,50"

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        response = self.client.post(
            reverse("import_price_list"),
            {
                "sheet": "HIRDAVAT",
                "title": "Liste 1",
                "action": "preview",
                "file": file_obj,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vida")
        self.assertNotContains(response, "<th>Fiyat</th>", html=False)
        self.assertNotContains(response, "10.50")

