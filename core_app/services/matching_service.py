from __future__ import annotations

from collections import defaultdict
from datetime import datetime

from django.http import HttpResponse

from core_app.models import PriceItem, PriceList, Stock
from core_app.services.data_quality_rules import get_issue_definition
from core_app.services.price_match_helpers import (
    build_price_item_key,
    build_price_item_lookup,
    build_stock_price_key,
    normalize_match_text,
    resolve_price_item_for_stock,
)


def get_active_price_list(request):
    active_id = request.session.get("active_price_list_id")
    if not active_id:
        return None, None
    return active_id, PriceList.objects.filter(id=active_id).first()


def build_name_index(rows):
    counts = defaultdict(int)
    groups = defaultdict(set)

    for row in rows:
        normalized_name = normalize_match_text(row.get("name"))
        group = (row.get("group") or "").strip()
        if not normalized_name:
            continue

        counts[normalized_name] += 1
        if group:
            groups[normalized_name].add(group)

    return counts, groups


def build_price_match_data(active_price_list_id: int):
    priceitems = list(
        PriceItem.objects.filter(price_list_id=active_price_list_id)
        .values("id", "name", "group", "price")
    )

    exact_price, name_only_price, groups_by_name = build_price_item_lookup(priceitems)
    stocks = list(Stock.objects.all().values("id", "name", "sku", "quantity", "unit", "category"))

    stock_exact_keys = set()
    stock_name_only_matches = set()
    stocks_without_price = []
    suspects = []

    for stock in stocks:
        normalized_name = normalize_match_text(stock.get("name"))
        sku_raw = (stock.get("sku") or "").strip()

        possible_groups = groups_by_name.get(normalized_name)
        groups_list = sorted(list(possible_groups)) if possible_groups else []
        suggested_group = groups_list[0] if len(groups_list) == 1 else ""

        price_item, match_type = resolve_price_item_for_stock(stock, exact_price, name_only_price)

        if sku_raw and possible_groups and match_type != "exact":
            issue = get_issue_definition("sku_group_conflict")
            suspects.append({
                "id": stock["id"],
                "name": stock.get("name"),
                "sku": stock.get("sku"),
                "category": stock.get("category"),
                "unit": stock.get("unit"),
                "quantity": stock.get("quantity"),
                "possible_groups": ", ".join(groups_list[:20]),
                "suggested_group": suggested_group,
                "note": "İsim var ama SKU/Grup uyuşmuyor",
                "issue": {"code": "sku_group_conflict", **issue},
            })

        if (not sku_raw) and possible_groups:
            issue = get_issue_definition("name_only_match")
            suspects.append({
                "id": stock["id"],
                "name": stock.get("name"),
                "sku": stock.get("sku"),
                "category": stock.get("category"),
                "unit": stock.get("unit"),
                "quantity": stock.get("quantity"),
                "possible_groups": ", ".join(groups_list[:20]),
                "suggested_group": suggested_group,
                "note": "SKU boş, fiyat listesinde grup var",
                "issue": {"code": "name_only_match", **issue},
            })

        if match_type == "exact":
            stock_exact_keys.add(build_stock_price_key(stock.get("name"), stock.get("sku")))
            if not sku_raw:
                stock_name_only_matches.add(normalized_name)
        elif match_type == "name_only":
            stock_name_only_matches.add(normalized_name)
        else:
            issue = get_issue_definition("stock_without_price")
            stock["issue"] = {"code": "stock_without_price", **issue}
            stocks_without_price.append(stock)

    priceitems_without_stock = []
    for price_item in priceitems:
        normalized_name = normalize_match_text(price_item.get("name"))
        raw_group = (price_item.get("group") or "").strip()
        price_key = build_price_item_key(price_item.get("name"), raw_group)

        if raw_group:
            if price_key not in stock_exact_keys:
                issue = get_issue_definition("price_without_stock")
                price_item["issue"] = {"code": "price_without_stock", **issue}
                priceitems_without_stock.append(price_item)
        elif normalized_name not in stock_name_only_matches:
            issue = get_issue_definition("price_without_stock")
            price_item["issue"] = {"code": "price_without_stock", **issue}
            priceitems_without_stock.append(price_item)

    return {
        "counts": {
            "stocks_total": len(stocks),
            "priceitems_total": len(priceitems),
            "stocks_without_price": len(stocks_without_price),
            "priceitems_without_stock": len(priceitems_without_stock),
            "suspects": len(suspects),
        },
        "stocks_without_price": stocks_without_price,
        "priceitems_without_stock": priceitems_without_stock,
        "suspects": suspects,
    }


def apply_price_match_filters(request, data):
    tab = (request.GET.get("tab") or "suspects").strip()
    q = (request.GET.get("q") or "").strip()
    category = (request.GET.get("category") or "").strip()
    sku_empty = (request.GET.get("sku_empty") or "") == "1"
    suggested_only = (request.GET.get("suggested_only") or "") == "1"
    multi_only = (request.GET.get("multi_only") or "") == "1"
    severity = (request.GET.get("severity") or "").strip()
    issue_code = (request.GET.get("issue_code") or "").strip()
    match_type = (request.GET.get("match_type") or "").strip()

    if suggested_only and multi_only:
        suggested_only = False
        multi_only = False

    def contains(haystack, needle_cf):
        return needle_cf in (haystack or "").casefold()

    q_cf = q.casefold()
    stocks_without_price = data["stocks_without_price"]
    suspects = data["suspects"]
    priceitems_without_stock = data["priceitems_without_stock"]

    if category:
        stocks_without_price = [row for row in stocks_without_price if (row.get("category") or "") == category]
        suspects = [row for row in suspects if (row.get("category") or "") == category]

    if sku_empty:
        stocks_without_price = [row for row in stocks_without_price if not (row.get("sku") or "").strip()]
        suspects = [row for row in suspects if not (row.get("sku") or "").strip()]

    if q:
        stocks_without_price = [
            row for row in stocks_without_price
            if contains(row.get("name"), q_cf) or contains(row.get("sku"), q_cf) or contains(row.get("category"), q_cf)
        ]
        suspects = [
            row for row in suspects
            if contains(row.get("name"), q_cf) or contains(row.get("sku"), q_cf) or contains(row.get("possible_groups"), q_cf)
        ]
        priceitems_without_stock = [
            row for row in priceitems_without_stock
            if contains(row.get("name"), q_cf) or contains(row.get("group"), q_cf)
        ]

    if suggested_only:
        suspects = [row for row in suspects if (row.get("suggested_group") or "").strip()]
    elif multi_only:
        suspects = [row for row in suspects if not (row.get("suggested_group") or "").strip()]

    if severity:
        suspects = [row for row in suspects if row.get("issue", {}).get("severity") == severity]
        stocks_without_price = [row for row in stocks_without_price if row.get("issue", {}).get("severity") == severity]
        priceitems_without_stock = [row for row in priceitems_without_stock if row.get("issue", {}).get("severity") == severity]

    if issue_code:
        suspects = [row for row in suspects if row.get("issue", {}).get("code") == issue_code]
        stocks_without_price = [row for row in stocks_without_price if row.get("issue", {}).get("code") == issue_code]
        priceitems_without_stock = [row for row in priceitems_without_stock if row.get("issue", {}).get("code") == issue_code]

    if match_type == "name_only":
        suspects = [row for row in suspects if row.get("issue", {}).get("code") == "name_only_match"]
    elif match_type == "sku_conflict":
        suspects = [row for row in suspects if row.get("issue", {}).get("code") == "sku_group_conflict"]

    filtered_counts = {
        "stocks_without_price": len(stocks_without_price),
        "priceitems_without_stock": len(priceitems_without_stock),
        "suspects": len(suspects),
    }

    qs = request.GET.urlencode()
    current_url = request.path + (("?" + qs) if qs else "")

    return {
        "tab": tab,
        "q": q,
        "category": category,
        "sku_empty": sku_empty,
        "suggested_only": suggested_only,
        "multi_only": multi_only,
        "severity": severity,
        "issue_code": issue_code,
        "match_type": match_type,
        "stocks_without_price": stocks_without_price,
        "priceitems_without_stock": priceitems_without_stock,
        "suspects": suspects,
        "filtered_counts": filtered_counts,
        "current_url": current_url,
    }


def get_price_match_metadata(active_price_list_id: int):
    categories = list(
        Stock.objects.exclude(category__isnull=True)
        .exclude(category="")
        .values_list("category", flat=True)
        .distinct()
    )
    all_groups = list(
        PriceItem.objects.filter(price_list_id=active_price_list_id)
        .exclude(group__isnull=True)
        .exclude(group="")
        .values_list("group", flat=True)
        .distinct()
    )

    return {
        "categories": sorted(set(category.strip() for category in categories if category and category.strip())),
        "all_groups": sorted(set(group.strip() for group in all_groups if group and group.strip())),
    }


def apply_price_match_bulk_sku(active_price_list_id: int, stock_ids, overwrite: bool, mode: str, chosen_group: str):
    groups_by_name = {}
    if mode == "suggested":
        rows = (
            PriceItem.objects.filter(price_list_id=active_price_list_id)
            .exclude(name__isnull=True).exclude(name="")
            .exclude(group__isnull=True).exclude(group="")
            .values_list("name", "group")
        )
        for name, group in rows:
            normalized_name = normalize_match_text(name)
            clean_group = (group or "").strip()
            if normalized_name and clean_group:
                groups_by_name.setdefault(normalized_name, set()).add(clean_group)

    updated = 0
    skipped_multi = 0
    skipped_no_suggestion = 0
    skipped_not_overwritten = 0

    for stock in Stock.objects.filter(id__in=stock_ids).values("id", "name", "sku"):
        current = (stock.get("sku") or "").strip()
        if current and not overwrite:
            skipped_not_overwritten += 1
            continue

        if mode == "manual":
            new_sku = chosen_group[:50]
        else:
            normalized_name = normalize_match_text(stock.get("name"))
            groups = sorted(list(groups_by_name.get(normalized_name, set())))
            if len(groups) == 1:
                new_sku = groups[0][:50]
            elif len(groups) > 1:
                skipped_multi += 1
                continue
            else:
                skipped_no_suggestion += 1
                continue

        Stock.objects.filter(id=stock["id"]).update(sku=new_sku)
        updated += 1

    return {
        "updated": updated,
        "skipped_multi": skipped_multi,
        "skipped_no_suggestion": skipped_no_suggestion,
        "skipped_not_overwritten": skipped_not_overwritten,
    }


def build_price_match_export_response(active_price_list, data, filters):
    from openpyxl import Workbook

    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["active_price_list", getattr(active_price_list, "title", "")])
    summary.append(["stocks_total", data["counts"]["stocks_total"]])
    summary.append(["priceitems_total", data["counts"]["priceitems_total"]])
    summary.append(["stocks_without_price_total", data["counts"]["stocks_without_price"]])
    summary.append(["priceitems_without_stock_total", data["counts"]["priceitems_without_stock"]])
    summary.append(["suspects_total", data["counts"]["suspects"]])
    summary.append([])
    summary.append(["filters"])
    summary.append(["tab", filters["tab"]])
    summary.append(["q", filters["q"]])
    summary.append(["category", filters["category"]])
    summary.append(["sku_empty", str(filters["sku_empty"])])

    stocks_sheet = workbook.create_sheet("stocks_without_price")
    stocks_sheet.append(["id", "name", "sku", "category", "unit", "quantity"])
    for row in filters["stocks_without_price"]:
        stocks_sheet.append([row.get("id"), row.get("name"), row.get("sku"), row.get("category"), row.get("unit"), row.get("quantity")])

    priceitems_sheet = workbook.create_sheet("priceitems_without_stock")
    priceitems_sheet.append(["id", "name", "group", "price"])
    for row in filters["priceitems_without_stock"]:
        priceitems_sheet.append([row.get("id"), row.get("name"), row.get("group"), row.get("price")])

    suspects_sheet = workbook.create_sheet("suspects")
    suspects_sheet.append(["id", "name", "sku", "suggested_group", "possible_groups", "note", "category", "unit", "quantity"])
    for row in filters["suspects"]:
        suspects_sheet.append([
            row.get("id"),
            row.get("name"),
            row.get("sku"),
            row.get("suggested_group"),
            row.get("possible_groups"),
            row.get("note"),
            row.get("category"),
            row.get("unit"),
            row.get("quantity"),
        ])

    filename = f"price_match_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
