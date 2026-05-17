import datetime as dt
from dataclasses import dataclass
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook

from core_app.models import PriceItem, PriceList
from core_app.services.text import normalize_text


class PriceImportError(ValueError):
    pass


@dataclass
class ParsedPriceItem:
    name: str
    group: str
    price: Decimal


@dataclass
class PriceImportResult:
    price_list: PriceList
    created_count: int


@dataclass
class PriceImportPreview:
    sheet_name: str
    list_date: dt.date | None
    item_count: int
    sample_items: list[ParsedPriceItem]


IGNORED_PRICE_ROW_NAMES = {
    "toplam",
    "genel toplam",
    "nakit",
    "fiyat",
    "ana fiyat",
    "cinsi",
    "olcu",
    "ölçü",
}


def to_decimal(value) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    text = (
        text.replace("₺", "")
        .replace("TL", "")
        .replace("tl", "")
        .replace("TRY", "")
        .replace("try", "")
        .strip()
    )

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")

    cleaned = []
    for ch in text:
        if ch.isdigit() or ch in {".", "-"}:
            cleaned.append(ch)
    text = "".join(cleaned).strip()

    if not text or text in {"-", ".", "-.", ".-"}:
        return None

    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except Exception:
        return None


def find_sheet_name(workbook, sheet_input: str) -> str | None:
    wanted = normalize_text(sheet_input)
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name) == wanted:
            return sheet_name
    return None


def detect_list_date(worksheet):
    max_c = min(worksheet.max_column, 20)
    for row_idx in range(1, 11):
        for col_idx in range(1, max_c + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if isinstance(value, dt.datetime):
                return value.date()
            if isinstance(value, dt.date):
                return value
    return None


def is_hirdavat2_sheet(sheet_name: str) -> bool:
    return normalize_text(sheet_name) in {"hirdavat 2", "hirdavat2"}


def should_skip_price_row(name: str) -> bool:
    return normalize_text(name) in IGNORED_PRICE_ROW_NAMES


def detect_profile_measure_header_row(worksheet) -> int | None:
    max_row = min(worksheet.max_row, 10)
    for row_idx in range(1, max_row + 1):
        col1 = normalize_text(worksheet.cell(row_idx, 1).value)
        col2 = normalize_text(worksheet.cell(row_idx, 2).value)
        col3 = normalize_text(worksheet.cell(row_idx, 3).value)
        col4 = normalize_text(worksheet.cell(row_idx, 4).value)
        if col1 == "cinsi" and col2 in {"olcu", "olcu"} and col3 == "cinsi" and "fiyat" in col4:
            return row_idx
    return None


def parse_profile_measure_sheet(worksheet, header_row: int) -> list[ParsedPriceItem]:
    items: list[ParsedPriceItem] = []
    current_group = ""

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        left_name_cell = worksheet.cell(row_idx, 1).value
        right_name_cell = worksheet.cell(row_idx, 3).value
        price_cell = worksheet.cell(row_idx, 4).value

        if left_name_cell is None and right_name_cell is None and price_cell is None:
            continue

        price = to_decimal(price_cell)
        if price is None:
            continue

        for name_cell in (left_name_cell, right_name_cell):
            if isinstance(name_cell, (dt.date, dt.datetime)):
                continue

            name = str(name_cell).strip() if name_cell is not None else ""
            if not name or should_skip_price_row(name):
                continue

            items.append(
                ParsedPriceItem(
                    name=name,
                    group=current_group,
                    price=price,
                )
            )

    return items


def parse_hirdavat2_sheet(worksheet) -> list[ParsedPriceItem]:
    items: list[ParsedPriceItem] = []
    blocks = list(range(1, worksheet.max_column + 1, 5))
    current_group = {block: "" for block in blocks}

    for row in worksheet.iter_rows(values_only=True):
        for start in blocks:
            name_cell = row[start - 1] if len(row) >= start else None
            price_cell = row[start] if len(row) >= (start + 1) else None

            if name_cell is None and price_cell is None:
                continue

            if isinstance(name_cell, (dt.date, dt.datetime)):
                continue

            name = str(name_cell).strip() if name_cell is not None else ""
            if not name:
                continue
            if should_skip_price_row(name):
                continue

            group_marker = str(price_cell).strip().upper() if isinstance(price_cell, str) else ""
            if group_marker in {"FİYAT", "FIYAT"}:
                if name.upper() not in {"GENEL TOPLAM", "TOPLAM"}:
                    current_group[start] = name
                continue

            price = to_decimal(price_cell)
            if price is None:
                continue

            group = current_group.get(start, "").strip()
            if not group:
                continue

            items.append(ParsedPriceItem(name=name, group=group, price=price))

    return items


def parse_standard_sheet(worksheet) -> list[ParsedPriceItem]:
    profile_measure_header_row = detect_profile_measure_header_row(worksheet)
    if profile_measure_header_row is not None:
        return parse_profile_measure_sheet(worksheet, profile_measure_header_row)

    items: list[ParsedPriceItem] = []
    pairs = [(1, 2), (4, 5), (7, 8)]
    current_group = {index: "" for index in range(len(pairs))}

    for row_idx in range(1, worksheet.max_row + 1):
        for pair_index, (name_col, price_col) in enumerate(pairs):
            raw_name = worksheet.cell(row_idx, name_col).value
            raw_price = worksheet.cell(row_idx, price_col).value

            if raw_name is None:
                continue

            if isinstance(raw_name, (dt.date, dt.datetime)):
                continue

            name = str(raw_name).strip()
            if not name:
                continue
            if should_skip_price_row(name):
                continue

            price = to_decimal(raw_price)
            if price is None:
                if raw_price is None and len(name) <= 60:
                    current_group[pair_index] = name
                continue

            items.append(
                ParsedPriceItem(
                    name=name,
                    group=current_group[pair_index],
                    price=price,
                )
            )

    return items


def parse_price_items(worksheet, sheet_name: str) -> list[ParsedPriceItem]:
    if is_hirdavat2_sheet(sheet_name):
        return parse_hirdavat2_sheet(worksheet)
    return parse_standard_sheet(worksheet)


def preview_price_list_workbook(file_obj, sheet_input: str, title: str) -> PriceImportPreview:
    sheet_input = (sheet_input or "").strip()
    title = (title or "").strip()

    if not sheet_input or not title:
        raise PriceImportError("Sheet adı ve Liste adı boş olamaz.")

    workbook = load_workbook(file_obj, data_only=True)
    target_sheet = find_sheet_name(workbook, sheet_input)
    if not target_sheet:
        preview = ", ".join(workbook.sheetnames[:20])
        raise PriceImportError(f"Sheet bulunamadı: '{sheet_input}'. Mevcut: {preview}")

    worksheet = workbook[target_sheet]
    list_date = detect_list_date(worksheet)
    parsed_items = parse_price_items(worksheet, target_sheet)

    if not parsed_items:
        raise PriceImportError("Fiyat listesinde içe aktarılacak geçerli kalem bulunamadı.")

    return PriceImportPreview(
        sheet_name=target_sheet,
        list_date=list_date,
        item_count=len(parsed_items),
        sample_items=parsed_items[:50],
    )


@transaction.atomic
def import_price_list_workbook(file_obj, sheet_input: str, title: str) -> PriceImportResult:
    sheet_input = (sheet_input or "").strip()
    title = (title or "").strip()

    if not sheet_input or not title:
        raise PriceImportError("Sheet adı ve Liste adı boş olamaz.")

    workbook = load_workbook(file_obj, data_only=True)
    target_sheet = find_sheet_name(workbook, sheet_input)
    if not target_sheet:
        preview = ", ".join(workbook.sheetnames[:20])
        raise PriceImportError(f"Sheet bulunamadı: '{sheet_input}'. Mevcut: {preview}")

    worksheet = workbook[target_sheet]
    list_date = detect_list_date(worksheet)
    parsed_items = parse_price_items(worksheet, target_sheet)

    if not parsed_items:
        raise PriceImportError("Fiyat listesinde içe aktarılacak geçerli kalem bulunamadı.")

    price_list = PriceList.objects.create(
        title=title,
        sheet_name=target_sheet,
        list_date=list_date,
    )

    for item in parsed_items:
        PriceItem.objects.create(
            price_list=price_list,
            group=item.group,
            name=item.name,
            price=item.price,
        )

    return PriceImportResult(price_list=price_list, created_count=len(parsed_items))

