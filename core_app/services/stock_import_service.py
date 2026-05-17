from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from zipfile import BadZipFile

from openpyxl import load_workbook

from core_app.services.stock_identity import create_or_merge_stock

REQUIRED_HEADERS = {"name", "quantity"}
ALLOWED_HEADERS = {"name", "category", "unit", "sku", "quantity"}
UNIT_MAP = {"adet": "adet", "kg": "kg", "mt": "mt", "metre": "mt"}


class StockImportError(ValueError):
    pass


@dataclass
class StockImportSummary:
    created: int
    updated: int
    skipped: int
    total: int
    processed: int
    errors: int
    error_reasons: dict[str, int]


def _normalize_header(value):
    return str(value).strip().lower() if value else ""


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return ""
    return str(value).strip()


def _parse_quantity(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return None

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace(",", ".")
    try:
        number = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None

    if number != number.to_integral_value():
        return None
    return int(number)


def _parse_unit(value):
    text = _cell_to_text(value).lower()
    if not text:
        return "adet", None

    unit = UNIT_MAP.get(text)
    if unit:
        return unit, None
    return None, "unit geçersiz"


def _is_blank_row(row):
    return all(value is None or _cell_to_text(value) == "" for value in row)


def _build_reason_counts(invalid_rows):
    counter = Counter()
    for row in invalid_rows:
        counter[row["reason"]] += 1
    return dict(counter)


def parse_stock_workbook(file_obj):
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except (BadZipFile, KeyError, OSError, ValueError) as exc:
        raise StockImportError("Excel dosyası okunamadı veya dosya formatı bozuk.") from exc

    worksheet = workbook.active
    if worksheet.max_row < 1:
        raise StockImportError("Excel dosyasında başlık satırı bulunamadı.")

    raw_headers = [cell.value for cell in worksheet[1]]
    headers = [_normalize_header(header) for header in raw_headers]
    non_empty_headers = {header for header in headers if header}

    if not non_empty_headers:
        raise StockImportError("Başlık satırı boş görünüyor.")

    missing = sorted(REQUIRED_HEADERS - non_empty_headers)
    extra = sorted(non_empty_headers - ALLOWED_HEADERS)

    def col_index(key):
        return headers.index(key) if key in headers else None

    idx_name = col_index("name")
    idx_category = col_index("category")
    idx_unit = col_index("unit")
    idx_sku = col_index("sku")
    idx_quantity = col_index("quantity")

    rows = []
    valid_rows = []
    invalid_rows = []
    blank_rows = 0

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_blank_row(row):
            blank_rows += 1
            continue

        name = _cell_to_text(row[idx_name]) if idx_name is not None else ""
        category = _cell_to_text(row[idx_category]) if idx_category is not None else ""
        sku = _cell_to_text(row[idx_sku]) if idx_sku is not None else ""
        unit, unit_error = _parse_unit(row[idx_unit] if idx_unit is not None else None)
        quantity = _parse_quantity(row[idx_quantity]) if idx_quantity is not None else None

        status = "ok"
        reason = ""
        if not name:
            status = "error"
            reason = "name boş"
        elif quantity is None:
            status = "error"
            reason = "quantity geçersiz"
        elif unit_error:
            status = "error"
            reason = unit_error

        row_data = {
            "row": row_idx,
            "name": name,
            "category": category,
            "unit": unit or "",
            "sku": sku,
            "quantity": quantity,
            "status": status,
            "reason": reason,
        }
        rows.append(row_data)
        if status == "ok":
            valid_rows.append(row_data)
        else:
            invalid_rows.append(row_data)

    reason_counts = _build_reason_counts(invalid_rows)

    return {
        "sheet_name": worksheet.title,
        "headers": headers,
        "missing_headers": missing,
        "extra_headers": extra,
        "rows": rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "total_rows": len(rows) + blank_rows,
        "processed_rows": len(rows),
        "blank_rows": blank_rows,
        "skipped_rows": len(invalid_rows) + blank_rows,
        "error_count": len(invalid_rows),
        "error_reasons": reason_counts,
    }


def import_stock_rows(parsed):
    created = 0
    updated = 0

    for row in parsed["valid_rows"]:
        result = create_or_merge_stock(
            name=row["name"],
            category=row["category"],
            subgroup="",
            unit=row["unit"],
            sku=row["sku"],
            quantity=row["quantity"],
        )
        if result.created:
            created += 1
        else:
            updated += 1

    return StockImportSummary(
        created=created,
        updated=updated,
        skipped=parsed["skipped_rows"],
        total=parsed["total_rows"],
        processed=parsed["processed_rows"],
        errors=parsed["error_count"],
        error_reasons=parsed["error_reasons"],
    )
